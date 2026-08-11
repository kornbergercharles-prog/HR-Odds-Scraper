import json
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment

print("SCRIPT STARTED")

def clean(x):
    if not x:
        return ""
    return str(x).replace("\xa0", " ").replace(".", "").replace("\n", " ").strip().lower()

def implied_prob(odds):
    try:
        o = int(float(str(odds).strip()))
        if o < 0:
            return (-o) / ((-o) + 100)
        else:
            return 100 / (o + 100)
    except:
        return None

def fmt_prob(val):
    return f"{val:.1%}" if val is not None else ""

def fmt_odds(val):
    try:
        o = int(float(val))
        return f"+{o}" if o > 0 else str(o)
    except:
        return "—"

def avg_prob_to_odds(prob):
    """Convert average implied probability back to American odds format"""
    try:
        if prob is None or prob <= 0 or prob >= 1:
            return "—"
        if prob >= 0.5:
            odds = -round((prob / (1 - prob)) * 100)
        else:
            odds = round(((1 - prob) / prob) * 100)
        return f"+{odds}" if odds > 0 else str(odds)
    except:
        return "—"

# Parse DraftKings
def parse_dk():
    try:
        with open("all_responses.json", "r", encoding="utf-8") as f:
            raw = json.load(f)
        if not raw:
            print("DraftKings: no data")
            return pd.DataFrame(columns=["player", "dk_odds"])
        df = pd.DataFrame(raw)
        if "player" not in df.columns or "dk_odds" not in df.columns:
            print("DraftKings: missing columns")
            return pd.DataFrame(columns=["player", "dk_odds"])
        df = df.dropna(subset=["player", "dk_odds"])
        df = df.drop_duplicates(subset=["player"])
        print(f"DraftKings players: {len(df)}")
        return df
    except Exception as e:
        print(f"DraftKings parse error: {e}")
        return pd.DataFrame(columns=["player", "dk_odds"])

# Parse FanDuel + Bovada from Odds API
def parse_odds_api():
    try:
        with open("odds_api_responses.json", "r", encoding="utf-8") as f:
            raw = json.load(f)
        if not raw:
            print("Odds API: no data")
            return pd.DataFrame(columns=["player", "fd_odds", "bovada_odds"])
        df = pd.DataFrame(raw)

        fd = df[df["book_key"] == "fanduel"][["player", "odds"]].rename(columns={"odds": "fd_odds"})
        bov = df[df["book_key"] == "bovada"][["player", "odds"]].rename(columns={"odds": "bovada_odds"})

        fd = fd.drop_duplicates(subset=["player"])
        bov = bov.drop_duplicates(subset=["player"])

        print(f"FanDuel players: {len(fd)}")
        print(f"Bovada players: {len(bov)}")

        merged = fd.merge(bov, on="player", how="outer")
        return merged

    except Exception as e:
        print(f"Odds API parse error: {e}")
        return pd.DataFrame(columns=["player", "fd_odds", "bovada_odds"])

# Read watchlist
with open("watchlist.txt", "r", encoding="utf-8") as f:
    watchlist_raw = [line.strip() for line in f if line.strip()]
watchlist = set(clean(x) for x in watchlist_raw)

dk_df = parse_dk()
api_df = parse_odds_api()

# Merge all three sources
if len(dk_df) > 0 and len(api_df) > 0:
    df = dk_df.merge(api_df, on="player", how="outer")
elif len(dk_df) > 0:
    df = dk_df.copy()
    df["fd_odds"] = None
    df["bovada_odds"] = None
elif len(api_df) > 0:
    df = api_df.copy()
    df["dk_odds"] = None
else:
    df = pd.DataFrame(columns=["player", "dk_odds", "fd_odds", "bovada_odds"])

for col in ["dk_odds", "fd_odds", "bovada_odds"]:
    if col not in df.columns:
        df[col] = None

# Calculate implied probs and averages
df["dk_implied"] = df["dk_odds"].apply(implied_prob)
df["fd_implied"] = df["fd_odds"].apply(implied_prob)
df["bov_implied"] = df["bovada_odds"].apply(implied_prob)
df["avg_implied"] = df[["dk_implied", "fd_implied", "bov_implied"]].mean(axis=1)
df["avg_odds"] = df["avg_implied"].apply(avg_prob_to_odds)
df = df.sort_values("avg_implied", ascending=False, na_position="last")

# Build output — avg odds and avg prob directly after player name
out = pd.DataFrame({
    "Player":           df["player"],
    "Avg Odds":         df["avg_odds"],
    "Avg Impl. Prob":   df["avg_implied"].apply(fmt_prob),
    "DK Odds":          df["dk_odds"].apply(lambda x: fmt_odds(x) if pd.notna(x) else "—"),
    "DK Impl. Prob":    df["dk_implied"].apply(fmt_prob),
    "FanDuel Odds":     df["fd_odds"].apply(lambda x: fmt_odds(x) if pd.notna(x) else "—"),
    "FD Impl. Prob":    df["fd_implied"].apply(fmt_prob),
    "Bovada Odds":      df["bovada_odds"].apply(lambda x: fmt_odds(x) if pd.notna(x) else "—"),
    "Bov Impl. Prob":   df["bov_implied"].apply(fmt_prob),
})

# Styling
yellow      = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
dark_blue   = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
cell_font   = Font(name="Arial", size=10)
center      = Alignment(horizontal="center")

with pd.ExcelWriter("hr_odds.xlsx", engine="openpyxl") as writer:
    out.to_excel(writer, index=False, sheet_name="HR Odds")
    ws = writer.sheets["HR Odds"]

    for cell in ws[1]:
        cell.fill      = dark_blue
        cell.font      = header_font
        cell.alignment = center

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    matched = 0
    for row_idx in range(2, ws.max_row + 1):
        is_watch = clean(str(ws.cell(row=row_idx, column=1).value)) in watchlist
        if is_watch:
            matched += 1
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font      = cell_font
            cell.alignment = center
            if is_watch:
                cell.fill = yellow

print(f"Saved hr_odds.xlsx")
print(f"Watchlist matches: {matched}")
